import os
import time
import copy
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys

EXPORT_FILE_NAME = 'test.xlsx'


def setup_driver():
    chrome_options = Options()
    chrome_options.add_argument("--incognito")
    # chrome_options.add_argument("--headless")
    chrome_options.add_argument("--window-size=1920x1080")
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    return webdriver.Chrome(options=chrome_options, executable_path='./chromedriver.exe')


def read_one_algorithm(algorithm):
    link = algorithm.find_elements_by_tag_name('a')[0].get_attribute('href')
    return {
        'Name': algorithm.text,
        'Link': link,
    }


def read_one_part(part):
    algorithms = part.find_elements_by_tag_name('li')
    title = part.find_elements_by_tag_name('strong')
    data = []
    for algorithm in algorithms:
        item = read_one_algorithm(algorithm)
        data.append(item)
    return {
        'Title': title[0].text,
        'Data': data
    }


def read_one_article(article):
    parts = article.find_elements_by_tag_name('li')
    data = []
    for part in parts:
        if len(part.find_elements_by_tag_name('strong')) > 0:
            item = read_one_part(part)
            data.append(item)
    return {
        'Data': data
    }


driver = setup_driver()
url = r"https://cp-algorithms.com/"
driver.get(url)
time.sleep(1)
articles_name = driver.find_elements_by_tag_name('h3')
articles = driver.find_elements_by_tag_name('ul')
cnt = 0
wb = openpyxl.Workbook()
for article in articles:
    if len(article.find_elements_by_tag_name('strong')) > 0:
        result = read_one_article(article)
        name = str(articles_name[cnt].text)
        df = pd.DataFrame(result['Data'])
        # for i in range(len(df['Title'])):
        #     print(df['Title'][i])
        #     for j in range(len(df['Data'][i])):
        #         print(df['Data'][i][j]['Name'], df['Data'][i][j]['Link'])
        #     print('\n')
        wb.create_sheet(name)
        ws = wb[name]
        cnt += 1
        print(df)
        row = 1
        for i in range(len(df['Title'])):
            ws.cell(row=row, column=1, value=df['Title'][i])
            for j in range(len(df['Data'][i])):
                ws.cell(row=row, column=2, value=df['Data'][i][j]['Name'])
                ws.cell(row=row, column=3, value=df['Data'][i][j]['Link'])
                row += 1
        wb.save('test.xlsx')

driver.close()